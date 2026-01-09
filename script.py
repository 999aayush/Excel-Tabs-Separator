import sys
import os
import subprocess
import shutil

# --- 1. SETUP ENVIRONMENT (Auto-Venv & CWD) ---
# Get the absolute directory where this script is located
script_dir = os.path.dirname(os.path.abspath(__file__))

# FORCE the Current Working Directory to be the script's folder
os.chdir(script_dir)

# Define expected venv path (standard Windows layout)
venv_python = os.path.join(script_dir, ".venv", "Scripts", "python.exe")

# AUTO-ACTIVATE VENV:
# If .venv exists and we aren't currently using it, relaunch this script using the venv python.
if os.path.exists(venv_python) and os.path.normcase(sys.executable) != os.path.normcase(venv_python):
    print(f"--> Activating Venv: {venv_python}")
    # Pass all original arguments to the new process
    subprocess.check_call([venv_python] + sys.argv)
    sys.exit()

# Now imports are safe because we are definitely in the venv
import openpyxl

# --- 2. INPUT HANDLING ---
print("\n--- Excel Tab Separator (Preserves Formatting) ---")
print(f"Running from: {script_dir}")

# Check if file was passed as an argument (dragged onto script icon)
if len(sys.argv) > 1:
    raw_input = sys.argv[1]
else:
    print("Tip: You can drag and drop your file into this window.")
    print("Note: If drag & drop is blocked (Admin Mode), right-click file -> 'Copy as Path' -> Paste here.")
    raw_input = input("Paste path or drag file here: ")

# --- 3. PATH CLEANING ---
# Remove surrounding whitespace
cleaned_path = raw_input.strip()

# Remove ampersand usually added by PowerShell drag-and-drop
if cleaned_path.startswith("&"):
    cleaned_path = cleaned_path[1:].strip()

# Remove both single and double quotes
input_path = cleaned_path.strip('"').strip("'")

# Verify file exists
if not os.path.exists(input_path):
    print(f"\nERROR: The file was not found at:\n{input_path}")
    print("Check the path and try again.")
    input("Press Enter to exit...") 
    sys.exit()

# --- 4. SETUP OUTPUT FOLDER ---
# Get the directory where the INPUT FILE lives (not the script)
base_directory = os.path.dirname(input_path)
file_name_only = os.path.splitext(os.path.basename(input_path))[0]

# Folder name is just the file name
output_dir = os.path.join(base_directory, file_name_only)

if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    print(f"\nCreated output folder: {output_dir}")
else:
    print(f"\nUsing existing output folder: {output_dir}")

print(f"Analyzing file...")

# --- 5. MAIN EXECUTION ---

try:
    # Load sheet names (Fast read-only)
    temp_wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    sheet_names = temp_wb.sheetnames
    temp_wb.close()
except Exception as e:
    print(f"Error opening file: {e}")
    input("Press Enter to exit...")
    sys.exit()

total_sheets = len(sheet_names)
print(f"Found {total_sheets} tabs to process.")

for index, target_sheet in enumerate(sheet_names):
    print(f"[{index+1}/{total_sheets}] Processing: {target_sheet}...")
    
    # Reload fresh copy
    wb = openpyxl.load_workbook(input_path)
    
    # Delete unwanted sheets
    for sheet in wb.sheetnames:
        if sheet != target_sheet:
            del wb[sheet]
            
    # Clean filename
    safe_name = "".join([c for c in target_sheet if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
    
    # Save
    final_output_path = os.path.join(output_dir, f"{safe_name}.xlsx")
    wb.save(final_output_path)
    wb.close()

print("-" * 30)
print("SUCCESS! All tabs have been separated.")
print(f"Location: {output_dir}")
input("Press Enter to close...")